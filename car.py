import pandas as pd
import openpyxl

class CarDeduplicator:
    def __init__(self, file_path, ignore_columns=None):
        self.file_path = file_path
        self.ignore_columns = ignore_columns if ignore_columns else []
        self.df = pd.read_excel(file_path)

    def read_excel(self):
        self.df = pd.read_excel(self.file_path)
    
    def remove_duplicates(self):
        removecolumns = self.detect_color()
        columns_to_check = [col for col in self.df.columns if col not in removecolumns]
        self.df = self.df.drop_duplicates(subset=columns_to_check)
    
    def save_to_excel(self, output_path):
        self.df.to_excel(output_path, index=False)
    
    def deduplicate(self, output_path):
        self.read_excel()
        self.remove_duplicates()
        self.save_to_excel(output_path)
    
    def detect_color(self):
        return None


class ToyotaDeduplicator(CarDeduplicator):
    def __init__(self, file_path):
        self.file_path = file_path 
        ignore_columns = self.detect_color()
        super().__init__(file_path, ignore_columns)
    
    def detect_color(self):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active 

        red_columns = []
        for col in sheet.iter_cols(1, sheet.max_column):
            cell = col[0]
            if self.is_red(cell):
                red_columns.append(cell.value)

        return red_columns

    def is_red(self, cell):
        fill = cell.fill
        if fill.start_color.index is not None:
            color = fill.start_color
            if color.type == "rgb" and color.rgb == "FFFF0000":
                return True
            elif color.type == "indexed" and color.indexed == 10:
                return True
        return False



toyota_deduplicator = ToyotaDeduplicator('SCION.xlsx')
toyota_deduplicator.deduplicate('SCION_Cleaned.xlsx')

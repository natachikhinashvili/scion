import pandas as pd
import openpyxl

from cardeduplicator import CarDeduplicator

class ToyotaDeduplicator(CarDeduplicator):
    def __init__(self, file_path):
        self.file_path = file_path 
        ignore_columns = self.detect_color()
        super().__init__(file_path, ignore_columns)
    

    def remove_duplicates(self):
        removecolumns = self.detect_color()
        columns_to_check = [col for col in self.df.columns if col not in removecolumns]
        self.df = self.df.drop_duplicates(subset=columns_to_check)
    
    def detect_color(self):
        wb = openpyxl.load_workbook(self.file_path)
        sheet = wb.active 

        red_columns = []
        for col in sheet.iter_cols(1, sheet.max_column):
            cell = col[0]
            if self.is_red(cell):
                red_columns.append(cell.value)

        return red_columns

    def is_red(self, cell):
        fill = cell.fill
        if fill.start_color.index is not None:
            color = fill.start_color
            if color.type == "rgb" and color.rgb == "FFFF0000":
                return True
            elif color.type == "indexed" and color.indexed == 10:
                return True
        return False